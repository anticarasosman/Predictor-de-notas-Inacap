from abc import ABC, abstractmethod
from openpyxl.utils import get_column_letter

class Sheet(ABC):

    @abstractmethod
    def add_sheet(self, workbook, data: dict):
        pass

    def auto_adjust_column_widths(self, worksheet, min_width: int = 12, max_width: int = 40, padding: int = 2) -> None:
        """Ajusta el ancho de columnas según el contenido y encabezados."""
        column_widths = {}

        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                value_length = len(str(cell.value))
                column_widths[cell.column] = max(column_widths.get(cell.column, 0), value_length)

        for col_idx, max_len in column_widths.items():
            adjusted_width = min(max(max_len + padding, min_width), max_width)
            worksheet.column_dimensions[get_column_letter(col_idx)].width = adjusted_width