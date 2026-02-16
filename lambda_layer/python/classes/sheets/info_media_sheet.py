from classes.sheets.sheet import Sheet
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl import Workbook

class InfoMediaSheet(Sheet):

    def add_sheet(self, wb, student_data: dict):
        student = student_data['student']
        ws = wb.create_sheet("Información Media")
        #Titulo
        ws["A1"] = "Notas Media Estudiante"
        ws["A1"].font = Font(size=14, bold=True)
        ws.merge_cells("A1:D1")

        row = 3

        # Datos del estudiante
        fields = [
            ('RUT', 'rut'),
            ('Nombre', 'nombre'),
            ('Promedio Matematica Media', 'promedio_media_matematica'),
            ('Promedio Lenguaje Media', 'promedio_media_lenguaje'),
            ('Promedio Inglés Media', 'promedio_media_ingles')
        ]

        for label, key, in fields:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = student.get(key, 'N/A')
            row += 1

        self.auto_adjust_column_widths(ws)