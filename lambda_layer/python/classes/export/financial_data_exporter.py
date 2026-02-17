from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path
from datetime import datetime


class FinancialDataExporter:
    """
    Exporta datos financieros de estudiantes con deuda.
    Incluye: RUT, Nombre, Programa, Deuda Total, Compromisos, Porcentaje de Morosidad, etc.
    """

    def __init__(self, db_connection):
        self.db_connection = db_connection

    def export_financial_data(self, output_dir: Path) -> str:
        """
        Exporta datos financieros de estudiantes con deuda.
        
        Args:
            output_dir: Directorio donde guardar el archivo Excel
            
        Returns:
            str: Ruta del archivo Excel generado
        """
        # Obtener datos financieros
        data = self._get_financial_data()
        
        if not data:
            raise ValueError("No hay estudiantes con deuda registrados")
        
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Data Financiera"
        
        # Crear hoja con datos
        self._create_financial_sheet(ws, data)
        
        # Guardar archivo
        file_path = self._save_workbook(wb, output_dir)
        
        return file_path

    def _get_financial_data(self) -> list:
        """
        Obtiene datos financieros de estudiantes con deuda > 0.
        
        Returns:
            list: Lista de diccionarios con información financiera
        """
        cursor = self.db_connection.cursor(dictionary=True)
        
        try:
            query = """
                SELECT 
                    e.rut,
                    COALESCE(e.nombre, '') AS nombre,
                    COALESCE(e.programa_estudio, '') AS programa_estudio,
                    COALESCE(rfe.deuda_total, 0) AS deuda_total,
                    COALESCE(rfe.monto_compromiso_matricula, 0) AS monto_compromiso_matricula,
                    COALESCE(rfe.monto_compromiso_colegiaturas, 0) AS monto_compromiso_colegiaturas,
                    COALESCE(rfe.deuda_matriculas, 0) AS deuda_matriculas,
                    COALESCE(rfe.deuda_colegiaturas, 0) AS deuda_colegiaturas,
                    COALESCE(rfe.otras_deudas, 0) AS otras_deudas,
                    COALESCE(rfe.cantidad_cuotas_pendientes_matriculas, 0) AS cuotas_pendientes_matricula,
                    COALESCE(rfe.cantidad_cuotas_pendientes_colegiaturas, 0) AS cuotas_pendientes_colegiatura
                FROM Reporte_financiero_estudiante rfe
                INNER JOIN Estudiante e ON e.rut = rfe.rut_estudiante
                WHERE rfe.deuda_total > 0
                ORDER BY rfe.deuda_total DESC
            """
            cursor.execute(query)
            
            data = []
            for row in cursor.fetchall():
                # Calcular porcentaje de morosidad
                total_compromisos = row['monto_compromiso_matricula'] + row['monto_compromiso_colegiaturas']
                
                if total_compromisos > 0:
                    porcentaje_morosidad = (row['deuda_total'] / total_compromisos) * 100
                else:
                    porcentaje_morosidad = 0
                
                row['total_compromisos'] = total_compromisos
                row['porcentaje_morosidad'] = round(porcentaje_morosidad, 2)
                
                data.append(row)
            
            return data
            
        finally:
            cursor.close()

    def _create_financial_sheet(self, ws, data: list):
        """
        Crea la hoja con datos financieros.
        """
        # Configurar encabezados
        headers = [
            'RUT',
            'Nombre',
            'Programa de Estudio',
            'Deuda Matrículas',
            'Deuda Colegiaturas',
            'Otras Deudas',
            'Deuda Total',
            'Compromiso Matrícula',
            'Compromiso Colegiaturas',
            'Total Compromisos',
            '% Morosidad',
            'Cuotas Pendientes (Matrícula)',
            'Cuotas Pendientes (Colegiatura)'
        ]
        
        # Estilo para encabezados
        header_fill = PatternFill(start_color="C4161C", end_color="C4161C", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Escribir encabezados
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
        
        # Formato para datos
        data_alignment = Alignment(horizontal="left", vertical="center")
        currency_format = '$#,##0'
        percentage_format = '0.00"%"'
        
        # Escribir datos
        for row_idx, row_data in enumerate(data, start=2):
            col_idx = 1
            
            # RUT
            cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get('rut', ''))
            cell.alignment = data_alignment
            cell.border = border
            col_idx += 1
            
            # Nombre
            cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get('nombre', ''))
            cell.alignment = data_alignment
            cell.border = border
            col_idx += 1
            
            # Programa
            cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get('programa_estudio', ''))
            cell.alignment = data_alignment
            cell.border = border
            col_idx += 1
            
            # Deuda Matrículas
            cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get('deuda_matriculas', 0))
            cell.number_format = currency_format
            cell.border = border
            col_idx += 1
            
            # Deuda Colegiaturas
            cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get('deuda_colegiaturas', 0))
            cell.number_format = currency_format
            cell.border = border
            col_idx += 1
            
            # Otras Deudas
            cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get('otras_deudas', 0))
            cell.number_format = currency_format
            cell.border = border
            col_idx += 1
            
            # Deuda Total
            cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get('deuda_total', 0))
            cell.number_format = currency_format
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
            cell.border = border
            col_idx += 1
            
            # Compromiso Matrícula
            cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get('monto_compromiso_matricula', 0))
            cell.number_format = currency_format
            cell.border = border
            col_idx += 1
            
            # Compromiso Colegiaturas
            cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get('monto_compromiso_colegiaturas', 0))
            cell.number_format = currency_format
            cell.border = border
            col_idx += 1
            
            # Total Compromisos
            cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get('total_compromisos', 0))
            cell.number_format = currency_format
            cell.border = border
            col_idx += 1
            
            # % Morosidad
            cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get('porcentaje_morosidad', 0))
            cell.number_format = percentage_format
            cell.font = Font(bold=True, color="C4161C")
            cell.border = border
            col_idx += 1
            
            # Cuotas Pendientes Matrícula
            cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get('cuotas_pendientes_matricula', 0))
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
            col_idx += 1
            
            # Cuotas Pendientes Colegiatura
            cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get('cuotas_pendientes_colegiatura', 0))
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        
        # Ajustar ancho de columnas
        column_widths = {
            'A': 15,   # RUT
            'B': 30,   # Nombre
            'C': 35,   # Programa
            'D': 18,   # Deuda Matrículas
            'E': 18,   # Deuda Colegiaturas
            'F': 15,   # Otras Deudas
            'G': 15,   # Deuda Total
            'H': 20,   # Compromiso Matrícula
            'I': 22,   # Compromiso Colegiaturas
            'J': 18,   # Total Compromisos
            'K': 14,   # % Morosidad
            'L': 25,   # Cuotas Pendientes Matrícula
            'M': 25    # Cuotas Pendientes Colegiatura
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Congelar primera fila
        ws.freeze_panes = 'A2'

    def _save_workbook(self, wb: Workbook, output_dir: Path) -> str:
        """
        Guarda el workbook en un archivo Excel.
        """
        # Crear directorio si no existe
        output_dir.mkdir(parents=True, exist_ok=True)
        
        # Nombre del archivo
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"Reporte_Financiero_Morosidad_{timestamp}.xlsx"
        
        file_path = output_dir / filename
        
        # Guardar
        wb.save(file_path)
        
        return str(file_path)
