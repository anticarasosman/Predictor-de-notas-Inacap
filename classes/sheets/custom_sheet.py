from classes.sheets.sheet import Sheet
from openpyxl.styles import Font, PatternFill, Alignment
from utils.dynamic_query_builder import build_query_for_custom_sheet

class CustomSheet(Sheet):
    def __init__(self, config: dict):
        self.name = config.get("name")
        self.tables = config.get("tables")

    def add_sheet(self, wb, student_data: dict, db_connection=None) -> None:
        """
        Agrega una hoja personalizada al workbook
        
        Args:
            wb: Workbook de openpyxl
            student_data: Datos del estudiante (incluye 'student' con RUT)
            db_connection: Conexión a BD (necesaria para queries de tablas personalizadas)
        """
        ws = wb.create_sheet(self.name)
        
        # Título
        ws["A1"] = f"Hoja Personalizada: {self.name}"
        ws["A1"].font = Font(size=14, bold=True)
        ws.merge_cells("A1:F1")
        
        row = 3
        
        # Iterar sobre cada tabla configurada
        for table_config in self.tables:
            table_name = table_config.get("table")
            columns = table_config.get("columns", [])
            
            if not columns:
                continue
            
            # Título de la tabla
            ws[f"A{row}"] = f"Tabla: {table_name}"
            ws[f"A{row}"].font = Font(bold=True, size=11)
            row += 1
            
            # Headers con colores
            header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for col_idx, column in enumerate(columns, 1):
                cell = ws.cell(row=row, column=col_idx, value=column)
                cell.fill = header_fill
                cell.font = header_font
            
            row += 1
            
            # Obtener datos de la BD
            try:
                if db_connection:
                    rut = student_data['student'].get('rut')
                    cursor = db_connection.cursor(dictionary=True)
                    
                    try:
                        # Construir y ejecutar query
                        query, params = build_query_for_custom_sheet(rut, table_name, columns)
                        cursor.execute(query, params)
                        data = cursor.fetchall()
                        
                        # Insertar filas de datos
                        for data_row in data:
                            for col_idx, column in enumerate(columns, 1):
                                value = data_row.get(column, "N/A")
                                ws.cell(row=row, column=col_idx, value=value)
                            row += 1
                    
                    finally:
                        cursor.close()
            
            except Exception as e:
                # Si hay error, mostrar mensaje en la celda
                ws[f"A{row}"] = f"Error: {str(e)}"
                row += 1
            
            row += 2  # Espacio entre tablas
        
        # Ajustar ancho de columnas
        for col_idx in range(1, 7):
            ws.column_dimensions[chr(64 + col_idx)].width = 18