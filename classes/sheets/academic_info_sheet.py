from classes.sheets.sheet import Sheet
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl import Workbook

class AcademicInfoSheet(Sheet):

    def add_sheet(self, wb, student_data: dict):
        semester_data = student_data["semesters"]
        ws = wb.create_sheet("Información Académica")
        header_fill = PatternFill(start_color = "2196F3", end_color="2196F3", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        row = 1

        for semester_obj in semester_data:
            semester = semester_obj['semester']
            subjects = semester_obj['subjects']

            #Titulo del semestre
            ws[f"A{row}"] = f"SEMESTRE: {semester['periodo_semestre']}"
            ws[f"A{row}"].font = Font(bold=True, size=12)
            ws.merge_cells(f"A{row}:F{row}")
            row += 2

            #Encabezados de las asignaturas
            headers = ['Código Asignatura', 'Docente', 'Notas Parciales', '% Asistencia', 'Riesgo']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font

            row += 1

            # Datos de asignaturas
            for subject in subjects:
                ws.cell(row=row, column=1, value=subject.get('codigo_asignatura', ''))
                ws.cell(row=row, column=2, value=subject.get('nombre_docente', ''))
                ws.cell(row=row, column=3, value=subject.get('notas_parciales', ''))
                ws.cell(row=row, column=4, value=subject.get('porcentaje_asistencia', ''))
                ws.cell(row=row, column=5, value="Sí" if subject.get('riesgo', False) else "No")
                row += 1

            row += 2  # Espacio entre semestres

        # Ajustar ancho de columnas
        self.auto_adjust_column_widths(ws)