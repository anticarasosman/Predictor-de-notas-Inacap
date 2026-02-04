from classes.sheets.sheet import Sheet
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl import Workbook

class GeneralInfoSheet(Sheet):

    def add_sheet(self, wb, student_data: dict):
        student = student_data['student']
        ws = wb.create_sheet("Información General")
        #Titulo
        ws["A1"] = "Información General del Estudiante"
        ws["A1"].font = Font(size=14, bold=True)
        ws.merge_cells("A1:D1")

        row = 3

        # Datos del estudiante
        fields = [
            ('RUT', 'rut'),
            ('Nombre', 'nombre'),
            ('Programa de Estudio', 'programa_estudio'),
            ('Tipo de Alumno', 'tipo_alumno'),
            ('Estado Matrícula', 'estado_matricula'),
            ('Terminal', 'terminal'),
            ('Tiene Gratuidad', 'tiene_gratuidad'),
            ('Deuda', 'deuda'),
            ('Nombre Apoderado', 'nombre_apoderado'),
        ]

        for label, key, in fields:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = student.get(key, 'N/A')
            row += 1

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 40