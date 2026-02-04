from classes.sheets.sheet import Sheet
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl import Workbook

class FinancialInfoSheet(Sheet):

    def add_sheet(self, wb, student_data: dict):
        financial_info = student_data["financial_info"]
        ws = wb.create_sheet("Información Financiera")

        #Titutlo
        ws["A1"] = "Información Financiera del Estudiante"
        ws["A1"].font = Font(size=14, bold=True)
        ws.merge_cells("A1:D1")

        row = 3

        fields = [
            ('Cuotas Pendientes Matrículas', 'cantidad_cuotas_pendientes_matriculas'),
            ('Cuotas Pendientes Colegiaturas', 'cantidad_cuotas_pendientes_colegiaturas'),
            ('Deuda Matrículas', 'deuda_matriculas'),
            ('Deuda Colegiaturas', 'deuda_colegiaturas'),
            ('Otras Deudas', 'otras_deudas'),
            ('Deuda Total', 'deuda_total'),
        ]

        for label, key in fields:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = financial_info.get(key, 'N/A')
            row += 1
        
        # Ajustar ancho de columnas
        self.auto_adjust_column_widths(ws)
        self.export_dir = "exports"