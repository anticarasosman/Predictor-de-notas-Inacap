import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import os
from datetime import datetime
from pathlib import Path

class StudentExporter:

    def __init__(self, db_connection):
        self.db_connection = db_connection
    
    def export_student_by_rut(self, rut: str, output_dir: Path) -> str:
        student_data = self._get_student_data(rut)
        
        if not student_data:
            raise ValueError(f"Estudiante con RUT {rut} no encontrado en la base de datos")
        
        file_path = self._create_excel(student_data, output_dir, rut)
        
        return file_path
    
    def _get_student_data(self, rut: str) -> dict:
        cursor = self.db_connection.cursor(dictionary=True)

        try:
            # 1. Datos generales del estudiante
            query = "SELECT * FROM estudiante WHERE rut = %s"
            cursor.execute(query, (rut,))
            student = cursor.fetchone()

            if not student:
                return None
            
            # 2. Semestres del estudiante
            query = """
                SELECT * FROM estudiante_semestre
                WHERE rut_estudiante = %s
                ORDER BY periodo_semestre DESC
            """
            cursor.execute(query, (rut,))
            semesters = cursor.fetchall()

            # 3. Cursos por semestre
            semesters_with_subjects = []
            for semester in semesters:
                query = """
                    SELECT * FROM estudiante_asignatura
                    WHERE rut_estudiante = %s AND periodo_semestre = %s
                    ORDER BY codigo_asignatura
                """
                cursor.execute(query, (rut, semester['periodo_semestre']))
                subjects = cursor.fetchall()
                semesters_with_subjects.append({
                    'semester': semester,
                    'subjects': subjects
                })

            # 4. Informacion financiera
            query = "SELECT * FROM reporte_financiero_estudiante WHERE rut_estudiante = %s"
            cursor.execute(query, (rut,))
            financial_info = cursor.fetchone()

            return {
                "student": student,
                "semesters": semesters_with_subjects,
                "financial_info": financial_info
            }
        
        finally:
            cursor.close()
    
    def _create_excel(self, student_data: dict, output_dir: Path, rut: str) -> str:
        wb = Workbook()
        wb.remove(wb.active)

        student = student_data['student']

        # 1. Hoja de Información General
        ws_general = wb.create_sheet("Información General", 0)
        self._add_general_info_sheet(ws_general, student)

        # 2. Hojas de Semestres y Cursos
        ws_semesters = wb.create_sheet("Semestres y Asignaturas", 1)
        self._add_semesters_sheet(ws_semesters, student_data['semesters'])

        # 3. Hoja de Información Financiera (Si existe)
        if student_data['financial_info']:
            ws_financial = wb.create_sheet("Información Financiera", 2)
            self._add_financial_info_sheet(ws_financial, student_data['financial_info'])

        # Guardar archivo
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"Estudiante_{rut}_{timestamp}.xlsx"
        file_path = output_dir / filename
        wb.save(file_path)
        return str(file_path)
    
    def _add_general_info_sheet(self, ws, student: dict):
        #Titulo
        ws["A1"] = "Información General del Estudiante"
        ws["A1"].font = Font(size=14, bold=True)
        ws.merge_cells("A1:D1")

        row = 3

        # Datos del estudiante
        fields = [
            ('RUT', 'rut'),
            ('Nombre', 'nombre'),
            ('Programa de Estudio', 'programa_estudio'),
            ('Tipo de Alumno', 'tipo_alumno'),
            ('Estado Matrícula', 'estado_matricula'),
            ('Terminal', 'terminal'),
            ('Tiene Gratuidad', 'tiene_gratuidad'),
            ('Deuda', 'deuda'),
            ('Nombre Apoderado', 'nombre_apoderado'),
        ]

        for label, key, in fields:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = student.get(key, 'N/A')
            row += 1

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 40

    def _add_semesters_sheet(self, ws, semester_data: list):

        header_fill = PatternFill(start_color = "2196F3", end_color="2196F3", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        row = 1

        for semester_obj in semester_data:
            semester = semester_obj['semester']
            subjects = semester_obj['subjects']

            #Titutlo del semestre
            ws[f"A{row}"] = f"SEMESTRE: {semester['periodo_semestre']}"
            ws[f"A{row}"].font = Font(bold=True, size=12)
            ws.merge_cells(f"A{row}:F{row}")
            row += 2

            #Encabezados de las asignaturas
            headers = ['Código Asignatura', 'Docente', 'Notas Parciales', '% Asistencia', 'Riesgo']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font

            row += 1

            # Datos de asignaturas
            for subject in subjects:
                ws.cell(row=row, column=1, value=subject.get('codigo_asignatura', ''))
                ws.cell(row=row, column=2, value=subject.get('nombre_docente', ''))
                ws.cell(row=row, column=3, value=subject.get('notas_parciales', ''))
                ws.cell(row=row, column=4, value=subject.get('porcentaje_asistencia', ''))
                ws.cell(row=row, column=5, value="Sí" if subject.get('riesgo', False) else "No")
                row += 1

            row += 2  # Espacio entre semestres

        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 12

    def _add_financial_info_sheet(self, ws, financial_info: list) -> None:
        #Titutlo

        ws["A1"] = "Información Financiera del Estudiante"
        ws["A1"].font = Font(size=14, bold=True)
        ws.merge_cells("A1:D1")

        row = 3

        fields = [
            ('Cuotas Pendientes Matrículas', 'cantidad_cuotas_pendientes_matriculas'),
            ('Cuotas Pendientes Colegiaturas', 'cantidad_cuotas_pendientes_colegiaturas'),
            ('Deuda Matrículas', 'deuda_matriculas'),
            ('Deuda Colegiaturas', 'deuda_colegiaturas'),
            ('Otras Deudas', 'otras_deudas'),
            ('Deuda Total', 'deuda_total'),
        ]

        for label, key in fields:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = financial_info.get(key, 'N/A')
            row += 1
        
        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 20
        self.export_dir = "exports"