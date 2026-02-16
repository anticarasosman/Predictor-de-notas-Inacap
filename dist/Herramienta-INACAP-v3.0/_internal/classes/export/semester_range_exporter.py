from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path
from datetime import datetime


class SemesterRangeExporter:
    """
    Exporta datos de múltiples estudiantes agrupados por semestre.
    Cada hoja representa un semestre con formato:
    - RUT | Programa de Estudio | Ramo | Nota | Docente | Info Asignaturas Reprobadas
    """

    def __init__(self, db_connection):
        self.db_connection = db_connection

    def export_by_semester_range(self, periodo_inicio: str, periodo_fin: str, output_dir: Path) -> str:
        """
        Exporta datos de todos los estudiantes entre dos periodos.
        
        Args:
            periodo_inicio: Periodo inicial (ej: 'OTONO 2024', 'PRIMAVERA 2024')
            periodo_fin: Periodo final (ej: 'OTONO 2025', 'PRIMAVERA 2025')
            output_dir: Directorio donde guardar el archivo Excel
            
        Returns:
            str: Ruta del archivo Excel generado
        """
        # Obtener lista de semestres en el rango
        semestres = self._get_semestres_in_range(periodo_inicio, periodo_fin)
        
        if not semestres:
            raise ValueError(f"No se encontraron semestres entre {periodo_inicio} y {periodo_fin}")
        
        # Crear workbook
        wb = Workbook()
        wb.remove(wb.active)  # Eliminar hoja por defecto
        
        # Crear una hoja por semestre
        for semestre in semestres:
            data = self._get_semester_data(semestre)
            self._create_semester_sheet(wb, semestre, data)
        
        # Guardar archivo
        file_path = self._save_workbook(wb, output_dir, periodo_inicio, periodo_fin)
        
        return file_path

    def _get_semestres_in_range(self, periodo_inicio: str, periodo_fin: str) -> list:
        """
        Obtiene todos los semestres entre dos periodos (inclusive).
        """
        cursor = self.db_connection.cursor(dictionary=True)
        
        try:
            query = """
                SELECT DISTINCT periodo 
                FROM Semestre
                ORDER BY periodo
            """
            cursor.execute(query)
            all_semestres = [row['periodo'] for row in cursor.fetchall()]
            
            # Filtrar semestres en el rango
            if periodo_inicio not in all_semestres or periodo_fin not in all_semestres:
                # Si no existen los periodos exactos, retornar todos
                return all_semestres
            
            idx_inicio = all_semestres.index(periodo_inicio)
            idx_fin = all_semestres.index(periodo_fin)
            
            # Asegurar orden correcto
            if idx_inicio > idx_fin:
                idx_inicio, idx_fin = idx_fin, idx_inicio
            
            return all_semestres[idx_inicio:idx_fin + 1]
            
        finally:
            cursor.close()

    def _get_semester_data(self, periodo: str) -> list:
        """
        Obtiene todos los datos de estudiantes para un semestre específico.
        
        Returns:
            list: Lista de diccionarios con:
                - rut: RUT del estudiante
                - nombre: Nombre del estudiante
                - programa: Programa de estudio
                - ramo: Nombre del ramo
                - notas_parciales: Notas separadas por comas
                - docente: Nombre del docente
                - asignaturas_reprobadas_cuatro_veces
                - asignaturas_reprobadas_tres_veces
        """
        cursor = self.db_connection.cursor(dictionary=True)
        
        try:
            query = """
                SELECT 
                    e.rut,
                    COALESCE(e.nombre, '') AS nombre,
                    COALESCE(e.programa_estudio, '') AS programa_estudio,
                    COALESCE(a.nombre, 'Sin asignaturas registradas') AS ramo,
                    COALESCE(ea.notas_parciales, '') AS notas_parciales,
                    COALESCE(ea.nombre_docente, '') AS docente,
                    COALESCE(es.asignaturas_reprobadas_cuatro_veces, 0) AS asignaturas_reprobadas_cuatro_veces,
                    COALESCE(es.asignaturas_reprobadas_tres_veces, 0) AS asignaturas_reprobadas_tres_veces
                FROM Estudiante_Semestre es
                INNER JOIN Estudiante e ON e.rut = es.rut_estudiante
                LEFT JOIN Estudiante_Asignatura ea ON ea.rut_estudiante = es.rut_estudiante 
                    AND ea.periodo_semestre = es.periodo_semestre
                LEFT JOIN Asignatura a ON a.codigo_asignatura = ea.codigo_asignatura
                WHERE es.periodo_semestre = %s
                ORDER BY e.rut, a.nombre
            """
            cursor.execute(query, (periodo,))
            return cursor.fetchall()
            
        finally:
            cursor.close()

    def _create_semester_sheet(self, wb: Workbook, semestre: str, data: list):
        """
        Crea una hoja en el workbook para un semestre específico.
        """
        # Nombre de hoja válido (max 31 caracteres, sin caracteres especiales)
        sheet_name = semestre.replace('/', '-')[:31]
        ws = wb.create_sheet(title=sheet_name)
        
        # Determinar número máximo de notas para crear columnas dinámicas
        max_notas = 0
        for row_data in data:
            notas_str = row_data.get('notas_parciales', '')
            if notas_str:
                num_notas = len(notas_str.split(','))
                max_notas = max(max_notas, num_notas)
        
        # Si no hay notas, crear al menos 3 columnas
        if max_notas == 0:
            max_notas = 3
        
        # Configurar encabezados dinámicos
        headers = [
            'RUT',
            'Nombre',
            'Programa de Estudio',
            'Ramo'
        ]
        
        # Agregar columnas de notas (SN-1, SN-2, SN-3, ...)
        for i in range(1, max_notas + 1):
            headers.append(f'SN-{i}')
        
        headers.extend([
            'Docente',
            'Asignaturas Reprobadas (4 veces)',
            'Asignaturas Reprobadas (3 veces)'
        ])
        
        # Estilo para encabezados
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Escribir encabezados
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Escribir datos
        if not data:
            # Si no hay datos, agregar mensaje informativo
            total_cols = len(headers)
            ws.cell(row=2, column=1, value='No hay estudiantes registrados en este semestre')
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
            cell = ws.cell(row=2, column=1)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(italic=True, color="808080")
        else:
            for row_idx, row_data in enumerate(data, start=2):
                col_idx = 1
                
                # RUT
                ws.cell(row=row_idx, column=col_idx, value=row_data.get('rut', ''))
                col_idx += 1
                
                # Nombre
                ws.cell(row=row_idx, column=col_idx, value=row_data.get('nombre', ''))
                col_idx += 1
                
                # Programa
                ws.cell(row=row_idx, column=col_idx, value=row_data.get('programa_estudio', ''))
                col_idx += 1
                
                # Ramo
                ws.cell(row=row_idx, column=col_idx, value=row_data.get('ramo', ''))
                col_idx += 1
                
                # Notas parciales - dividir por cada columna
                notas_str = row_data.get('notas_parciales', '')
                if notas_str:
                    notas_list = [nota.strip() for nota in notas_str.split(',')]
                else:
                    notas_list = []
                
                # Escribir cada nota en su columna correspondiente
                for i in range(max_notas):
                    if i < len(notas_list):
                        ws.cell(row=row_idx, column=col_idx, value=notas_list[i])
                    else:
                        ws.cell(row=row_idx, column=col_idx, value='')
                    col_idx += 1
                
                # Docente
                ws.cell(row=row_idx, column=col_idx, value=row_data.get('docente', ''))
                col_idx += 1
                
                # Asignaturas reprobadas 4 veces
                ws.cell(row=row_idx, column=col_idx, value=row_data.get('asignaturas_reprobadas_cuatro_veces', ''))
                col_idx += 1
                
                # Asignaturas reprobadas 3 veces
                ws.cell(row=row_idx, column=col_idx, value=row_data.get('asignaturas_reprobadas_tres_veces', ''))
        
        # Ajustar ancho de columnas dinámicamente
        ws.column_dimensions['A'].width = 15  # RUT
        ws.column_dimensions['B'].width = 30  # Nombre
        ws.column_dimensions['C'].width = 40  # Programa
        ws.column_dimensions['D'].width = 35  # Ramo
        
        # Columnas de notas (SN-1, SN-2, etc.)
        for i in range(max_notas):
            col_letter = chr(ord('E') + i)  # E, F, G, ...
            ws.column_dimensions[col_letter].width = 10
        
        # Docente y reprobadas
        docente_col = chr(ord('E') + max_notas)
        ws.column_dimensions[docente_col].width = 30
        
        reprobadas_4_col = chr(ord('E') + max_notas + 1)
        ws.column_dimensions[reprobadas_4_col].width = 25
        
        reprobadas_3_col = chr(ord('E') + max_notas + 2)
        ws.column_dimensions[reprobadas_3_col].width = 25
        
        # Congelar primera fila
        ws.freeze_panes = 'A2'

    def _save_workbook(self, wb: Workbook, output_dir: Path, periodo_inicio: str, periodo_fin: str) -> str:
        """
        Guarda el workbook en un archivo Excel.
        """
        # Crear directorio si no existe
        output_dir.mkdir(parents=True, exist_ok=True)
        
        # Nombre del archivo
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        inicio_clean = periodo_inicio.replace(' ', '_')
        fin_clean = periodo_fin.replace(' ', '_')
        filename = f"Reporte_Semestral_{inicio_clean}_a_{fin_clean}_{timestamp}.xlsx"
        
        file_path = output_dir / filename
        
        # Guardar
        wb.save(file_path)
        
        return str(file_path)
